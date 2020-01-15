import requests
from bs4 import BeautifulSoup
import openpyxl

# wb = openpyxl.Workbook()

# sheet = wb.active
# sheet['A1'] = 'hello world!'
# sheet.cell(row = 3, column = 3).value = 'BYE!'
#
# subject = ['Python', 'Java', 'HTML', 'JavaScript']
# sheet.append(subject)

# sheet1 = wb.active
# sheet1.title = '1st sheet'
# sheet2 = wb.create_sheet('2nd sheet')
#
# for  i in range(1,10):
#     sheet1.cell(row=i, column = 1).value = i
#     sheet2.cell(row=1, column = i).value = i
#
# wb.save('test.xlsx')

wb = openpyxl.Workbook()

sheet = wb.active

sheet.append(['제목', '채널명', '재생 수', '좋아요 수'])

raw = requests.get('https://tv.naver.com/r')

html = BeautifulSoup(raw.text, 'html.parser')

clips = html.select("div.inner")

for clip in clips :
    title = clip.select_one("dt.title").text.strip()
    chn = clip.select_one('dd.chn').text.strip()
    hit = clip.select_one("span.hit").text.strip().replace( ','  , '' )
    hit = hit[4:]

    like = clip.select_one("span.like").text.strip().replace(',', '')
    like = like[5:]

    sheet.append([title, chn, int(hit), int(like)])

wb.save('naver.xlsx')